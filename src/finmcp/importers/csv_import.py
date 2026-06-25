from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import date, datetime, timezone

from sqlalchemy.orm import Session

from finmcp.db import models

# Heurística de cabeceras (CaixaBank y exports comunes, en ES/EN).
_DATE_KEYS = (
    "f. operación",
    "f. operacion",
    "f.operación",
    "f.operacion",
    "fecha operación",
    "fecha operacion",
    "fecha contable",
    "fecha valor",
    "f. valor",
    "fecha",
    "date",
)
# Descripción: en CaixaBank el comercio está en "Concepto complementario 1".
_DESC_KEYS = (
    "complementario 1",
    "concepto propio",
    "concepto",
    "descripcion",
    "descripción",
    "detalle",
    "movimiento",
    "concept",
    "description",
)
_AMOUNT_KEYS = ("importe", "amount", "cantidad")
_INGRESO_KEYS = ("ingreso", "haber", "abono", "credit")
_GASTO_KEYS = ("gasto", "debe", "cargo", "debit")
_ACCOUNT_KEYS = ("número de cuenta", "numero de cuenta", "cuenta", "iban")


def _norm(s: str) -> str:
    return (s or "").strip().lower()


def parse_amount(raw: str) -> float | None:
    """'1.234,56' -> 1234.56 · '-45,00' -> -45.0 · '1234.56' -> 1234.56"""
    s = (raw or "").strip().replace("€", "").replace(" ", "").replace("\xa0", "")
    if not s:
        return None
    if "," in s and "." in s:  # formato ES: punto miles, coma decimal
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(raw: str) -> datetime | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _find_col(headers: list[str], keys: tuple[str, ...]) -> int | None:
    norm = [_norm(h) for h in headers]
    for key in keys:  # claves más específicas primero
        for i, h in enumerate(norm):
            if key in h:
                return i
    return None


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def parse_rows(text: str, delimiter: str | None = None) -> list[dict]:
    """Extrae movimientos de un CSV/TXT delimitado."""
    text = text.lstrip("﻿")  # quitar BOM
    if delimiter is None:
        sample = text.splitlines()[0] if text.strip() else ""
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    table = [r for r in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return _rows_from_table(table)


def parse_xlsx(path: str) -> list[dict]:
    """Extrae movimientos de un Excel moderno (.xlsx)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    table = [
        [_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)
    ]
    return _rows_from_table(table)


def parse_xls(path: str) -> list[dict]:
    """Extrae movimientos de un Excel antiguo (.xls), como el de CaixaBank."""
    import xlrd

    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    table = [
        [_cell_to_str(sh.cell_value(r, c)) for c in range(sh.ncols)]
        for r in range(sh.nrows)
    ]
    return _rows_from_table(table)


def _rows_from_table(table: list[list[str]]) -> list[dict]:
    """Localiza la cabecera (saltando preámbulos) y extrae las filas válidas.

    Soporta importe en una columna con signo o en columnas Ingreso/Gasto
    separadas (CaixaBank), y una columna opcional de número de cuenta.
    """
    header_idx = None
    for i, row in enumerate(table):
        has_date = _find_col(row, _DATE_KEYS) is not None
        has_amount = (
            _find_col(row, _AMOUNT_KEYS) is not None
            or _find_col(row, _INGRESO_KEYS) is not None
            or _find_col(row, _GASTO_KEYS) is not None
        )
        if has_date and has_amount:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "No encuentro las cabeceras (fecha/importe) en el fichero. "
            "Revisa el formato o el delimitador (--delimiter)."
        )

    h = table[header_idx]
    di = _find_col(h, _DATE_KEYS)
    ci = _find_col(h, _DESC_KEYS)
    ai = _find_col(h, _AMOUNT_KEYS)
    ii = _find_col(h, _INGRESO_KEYS)
    gi = _find_col(h, _GASTO_KEYS)
    acci = _find_col(h, _ACCOUNT_KEYS)

    def cell(row: list[str], idx: int | None) -> str:
        return row[idx].strip() if (idx is not None and idx < len(row)) else ""

    out: list[dict] = []
    for row in table[header_idx + 1 :]:
        d = parse_date(cell(row, di))
        if d is None:
            continue
        if ai is not None:
            amount = parse_amount(cell(row, ai))
        else:
            ing = parse_amount(cell(row, ii)) or 0.0
            gas = parse_amount(cell(row, gi)) or 0.0
            amount = ing - gas if (ing or gas) else None
        if amount is None or amount == 0:
            continue
        out.append(
            {
                "date": d,
                "amount": amount,
                "description": cell(row, ci),
                "account": cell(row, acci) or None,
            }
        )
    return out


def _digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def import_transactions(
    session: Session, rows: list[dict], account: models.Account | None = None
) -> tuple[int, int]:
    """Inserta movimientos deduplicando por (cuenta, día, importe, tipo).

    Cada fila se asigna a su cuenta: por el número de cuenta del fichero si lo
    trae (multi-cuenta), o a `account` en su defecto. Devuelve (importadas, saltadas).
    """
    accounts = session.query(models.Account).all()

    seen: dict[str, set] = {}

    def seen_for(acc: models.Account) -> set:
        if acc.id not in seen:
            existing = (
                session.query(models.Transaction)
                .filter(models.Transaction.account_id == acc.id)
                .all()
            )
            seen[acc.id] = {
                (t.booked_at.date().isoformat(), round(t.amount, 2), t.type)
                for t in existing
            }
        return seen[acc.id]

    def resolve(hint: str | None) -> models.Account | None:
        if hint:
            d = _digits(hint)
            match = next(
                (a for a in accounts if a.iban and _digits(a.iban).endswith(d)),
                None,
            )
            if match:
                return match
        return account

    added = skipped = 0
    for r in rows:
        acc = resolve(r.get("account"))
        if acc is None:
            skipped += 1
            continue
        side = "debit" if r["amount"] < 0 else "credit"
        amount = abs(r["amount"])
        key = (r["date"].date().isoformat(), round(amount, 2), side)
        bucket = seen_for(acc)
        if key in bucket:
            skipped += 1
            continue
        tid = "csv_" + hashlib.sha1(
            f"{acc.id}|{key}|{r['description']}".encode()
        ).hexdigest()[:24]
        if session.get(models.Transaction, tid) is not None:
            skipped += 1
            continue
        session.add(
            models.Transaction(
                id=tid,
                account_id=acc.id,
                amount=amount,
                currency=acc.currency,
                booked_at=r["date"],
                description=r["description"],
                type=side,
                merchant_name=r["description"] or None,
                provider_category=None,
                raw_json={"source": "csv"},
            )
        )
        bucket.add(key)
        added += 1
    session.commit()
    return added, skipped
